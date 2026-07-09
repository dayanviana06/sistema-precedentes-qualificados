# -*- coding: utf-8 -*-
"""Informativos de Jurisprudência do STF e do STJ (a partir de 2020).

STF — 'Planilha do Informativo STF' (arquivo estruturado oficial com todas
  as edições, divulgado pelo próprio Tribunal). Como o portal do STF bloqueia
  robôs, o arquivo entra pelo padrão local: baixe-o uma vez e salve como
  'stf-informativos.xlsx' na raiz do repositório; renove quando desejar.

STJ — duas vias: (a) arquivo local 'stj-informativos.csv' exportado da
  pesquisa oficial de informativos; e/ou (b) Espelhos de Acórdãos do Portal
  de Dados Abertos (CSV, licença CC-BY, automatizável) — ative preenchendo
  STJ_ESPELHOS_CSV em fontes.py com os recursos desejados.

Item de informativo NÃO é precedente vinculante por si: entra como espécie
'INF', marcado 'Consulta'. Disciplina de zero invenção: campo ilegível vira
NAO_LOCALIZADO e nunca sobrescreve dado anterior."""
import csv, io, os, re, unicodedata
from fontes import NAO_LOCALIZADO
from util import baixar, log

ANO_MINIMO = 2020
ARQ_STF = "stf-informativos.xlsx"
ARQ_STJ = "stj-informativos.csv"

try:
    from fontes import STJ_ESPELHOS_CSV
except ImportError:
    STJ_ESPELHOS_CSV = []

def _norm(k):
    k = "".join(c for c in unicodedata.normalize("NFD", str(k or ""))
                if unicodedata.category(c) != "Mn")
    return "".join(ch for ch in k.strip().lower() if ch.isalnum())

def _pega(r, *chaves):
    for c in chaves:
        if r.get(c):
            return str(r[c]).strip()
    return ""

def _ano(*textos):
    for t in textos:
        m = re.search(r"\b(20\d{2})\b", str(t or ""))
        if m:
            return int(m.group(1))
    return None

def _slug_txt(s):
    s = "".join(c for c in unicodedata.normalize("NFD", s)
                if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9]+", "-", s.lower()).strip("-")[:60]

def _item(trib, num_inf, titulo, teor, ramo, processo, data, origem):
    if not teor or len(teor) < 25:
        teor = NAO_LOCALIZADO
    ident = f"inf-{trib.lower()}-{num_inf}-{_slug_txt(titulo or teor[:60])}"
    return {
        "id": ident,
        "tribunal": trib, "especie": "INF",
        "area": ramo or f"Informativos do {trib}",
        "obrigatorio": False,                 # informativo divulga; não vincula por si
        "validade": "Vigente",
        "situacao": (f"Informativo {num_inf}/{trib}"
                     + (f" · {data}" if data else "")
                     + (f" · {processo}" if processo else "")),
        "assuntos": [a for a in {ramo, f"Informativos {trib}"} if a],
        "numero": f"Informativo {num_inf}/{trib}",
        "fonte": origem,
        "titulo": (titulo or teor[:150])[:180],
        "resumo": teor[:400], "completo": teor,
        "precedentes": processo or "", "vinculos": [],
    }

def _linhas_xlsx(caminho):
    from openpyxl import load_workbook
    wb = load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    linhas = ws.iter_rows(values_only=True)
    cab = [_norm(c) for c in next(linhas)]
    for row in linhas:
        yield dict(zip(cab, ["" if v is None else v for v in row]))

def _linhas_csv(texto):
    sep = ";" if texto[:3000].count(";") > texto[:3000].count(",") else ","
    for row in csv.DictReader(io.StringIO(texto), delimiter=sep):
        yield {_norm(k): (v or "").strip() for k, v in row.items()}

def _mapear(r, trib, origem):
    num = re.sub(r"\D", "", _pega(r, "informativo", "numeroinformativo",
                                  "edicao", "numero")) or "0"
    titulo = _pega(r, "titulo", "tema", "assunto", "tese", "ementa")[:400]
    teor = _pega(r, "teor", "resumo", "teseouteor", "texto", "conteudo",
                 "tese", "ementa", "julgado")
    ramo = _pega(r, "ramododireito", "ramo", "materia", "area",
                 "classificacao")
    proc = _pega(r, "processo", "classe", "leadingcase", "referencia")
    data = _pega(r, "data", "datajulgamento", "datadejulgamento",
                 "datapublicacao")
    ano = _ano(data, num if len(num) == 4 else "", _pega(r, "ano"))
    if ano and ano < ANO_MINIMO:
        return None
    if num == "0" and not teor:
        return None
    return _item(trib, num, titulo, teor, ramo, proc, data, origem)

def coletar():
    itens = []
    # STF — planilha oficial local
    if os.path.exists(ARQ_STF):
        try:
            n0 = len(itens)
            for r in _linhas_xlsx(ARQ_STF):
                it = _mapear(r, "STF", "STF — Planilha oficial do Informativo")
                if it:
                    itens.append(it)
            log(f"Informativos STF: {len(itens)-n0} itens (edições ≥ {ANO_MINIMO}) da planilha oficial.")
        except Exception as e:
            log(f"Informativos STF: falha na planilha ({e}) — base anterior preservada.")
    else:
        log(f"Informativos STF: '{ARQ_STF}' ausente — baixe a planilha oficial do Informativo e salve na raiz (opcional).")
    # STJ — arquivo local
    if os.path.exists(ARQ_STJ):
        try:
            with open(ARQ_STJ, encoding="utf-8-sig", errors="replace") as f:
                texto = f.read()
            n0 = len(itens)
            for r in _linhas_csv(texto):
                it = _mapear(r, "STJ", "STJ — Informativo de Jurisprudência (exportação oficial)")
                if it:
                    itens.append(it)
            log(f"Informativos STJ: {len(itens)-n0} itens (≥ {ANO_MINIMO}) do arquivo local.")
        except Exception as e:
            log(f"Informativos STJ: falha no arquivo local ({e}).")
    # STJ — espelhos de acórdãos (dados abertos, automatizável)
    for url in STJ_ESPELHOS_CSV:
        try:
            n0 = len(itens)
            for r in _linhas_csv(baixar(url)):
                it = _mapear(r, "STJ", "STJ — Dados Abertos (Espelhos de Acórdãos)")
                if it:
                    itens.append(it)
            log(f"Informativos/Espelhos STJ: {len(itens)-n0} itens de {url.rsplit('/',1)[-1]}.")
        except Exception as e:
            log(f"Espelhos STJ: fonte indisponível ({e}) — base anterior preservada.")
    return itens
